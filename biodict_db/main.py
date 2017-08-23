import sqlite3
from openpyxl import load_workbook
from os import *
from os.path import *
import io


skipped_picture = [149, 235, 245]

def read_image(file_name):
    try:
        fin = io.open(file_name, "rb")
        img = fin.read()
        return img
    except IOError as e:
        print("Error %d: %s" % (e.args[0],e.args[1]))
        sys.exit(1)
    finally:
        if fin:
            fin.close()

def is_null_or_empty(s):
	if s is None or len(s)==0:
		return True
	else:
		return False

def get_pic_list():
	path = 'pic/'
	pic_list = [f for f in listdir(path)]
	print(len(pic_list))
	print(pic_list)

	filtered_list =  [v for i, v in enumerate(pic_list) if (i+1) % 2 == 1]
	print(len(filtered_list))
	return filtered_list

def generate_data():
	data_wb = load_workbook("ENTRI KAMUS-1.xlsx")
	ws = data_wb.active
	pic_list = get_pic_list()
	index = 0
	for i, pic_name in zip(range(3, 124), pic_list):
		print(pic_name)
		data = read_image('pic/%s' % pic_name)
		binary = sqlite3.Binary(data)

		bahasa = ws[('B%d' % i)].value if not is_null_or_empty(ws[('B%d' % i)].value) else ""
		english = ws[('C%d' % i)].value if not is_null_or_empty(ws[('C%d' % i)].value) else ""
		latin = ws[('D%d' % i)].value if not is_null_or_empty(ws[('D%d' % i)].value) else ""
		kingdom = ws[('F%d' % i)].value if not is_null_or_empty(ws[('F%d' % i)].value) else ""
		subkingdom = ws[('G%d' % i)].value if not is_null_or_empty(ws[('G%d' % i)].value) else ""
		superdivisio = ws[('H%d' % i)].value if not is_null_or_empty(ws[('H%d' % i)].value) else ""
		divisio = ws[('I%d' % i)].value if not is_null_or_empty(ws[('I%d' % i)].value) else ""
		kelas = ws[('J%d' % i)].value if not is_null_or_empty(ws[('J%d' % i)].value) else ""
		subkelas = ws[('K%d' % i)].value if not is_null_or_empty(ws[('K%d' % i)].value) else ""
		ordo = ws[('L%d' % i)].value if not is_null_or_empty(ws[('L%d' % i)].value) else ""
		famili = ws[('M%d' % i)].value if not is_null_or_empty(ws[('M%d' % i)].value) else ""
		genus = ws[('N%d' % i)].value if not is_null_or_empty(ws[('N%d' % i)].value) else ""
		spesies = ws[('O%d' % i)].value if not is_null_or_empty(ws[('O%d' % i)].value) else ""
		morfologi = ws[('P%d' % i)].value if not is_null_or_empty(ws[('P%d' % i)].value) else ""
		manfaat = ws[('Q%d' % i)].value if not is_null_or_empty(ws[('Q%d' % i)].value) else ""
		yield bahasa, english, latin, kingdom, subkingdom, superdivisio, divisio, kelas, subkelas, ordo, famili, genus, spesies, morfologi, manfaat, binary

conn = sqlite3.connect("dict.db")
cur = conn.cursor()
cur.execute('''DROP TABLE IF EXISTS DictData''')
cur.execute('''CREATE TABLE `DictData` (
	`_id`	INTEGER NOT NULL UNIQUE,
	`NamaIndonesia`	TEXT NOT NULL,
	`NamaInggris`	TEXT,
	`NamaIlmiah`	TEXT,
	`Kingdom`	TEXT,
	`SubKingdom`	TEXT,
	`SuperDivisio`	TEXT,
	`Divisio`	TEXT,
	`Kelas`	TEXT,
	`SubKelas`	TEXT,
	`Ordo`	TEXT,
	`Famili`	TEXT,
	`Genus`	TEXT,
	`Spesies`	TEXT,
	`Morfologi`	TEXT,
	`Manfaat`	TEXT,
	`Gambar`	BLOB,
	PRIMARY KEY(`_id`)
);''')

for bahasa, english, latin, kingdom, subkingdom, superdivisio, divisio, kelas, subkelas, ordo, famili, genus, spesies, morfologi, manfaat, binary in generate_data():
	cur.execute('INSERT INTO DictData(NamaIndonesia, NamaInggris, NamaIlmiah, Kingdom, SubKingdom, SuperDivisio, Divisio, Kelas, SubKelas, Ordo, Famili, Genus, Spesies, Morfologi, Manfaat, Gambar) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', (bahasa, english, latin, kingdom, subkingdom, superdivisio, divisio, kelas, subkelas, ordo, famili, genus, spesies, morfologi, manfaat, binary))

conn.commit()

